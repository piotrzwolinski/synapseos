"""Generate a test Excel file with Camfil competitor products for cross-reference testing."""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path

OUTPUT_PATH = Path(__file__).parent / "camfil_test_order.xlsx"

ROWS = [
    ("Hi-Flo ES ePM1 65% 592x592x600", 12, "AHU 01"),
    ("Hi-Flo XLS ePM1 80% 592x592x635", 8, "AHU 02"),
    ("Opakfil GT ePM1 80% 592x592x292", 6, "Roof unit"),
    ("30/30 ePM10 50% 592x592x48", 24, "Pre-filter"),
    ("Absolute XH H13 610x610x292", 4, "Cleanroom"),
    ("CamCube HF-S", 2, "Housing"),
    ("Hi-Flo ES ePM2.5 65% 287x592x600", 10, "AHU 03 - half size"),
]

HEADERS = ["Product", "Quantity", "Notes"]


def create_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Camfil Order"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    data_font = Font(name="Calibri", size=11)
    for row_idx, (product, quantity, notes) in enumerate(ROWS, start=2):
        for col_idx, value in enumerate([product, quantity, notes], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 25

    return wb


def main() -> None:
    wb = create_workbook()
    wb.save(str(OUTPUT_PATH))
    print(f"Created: {OUTPUT_PATH}")
    print(f"Rows: {len(ROWS)} data rows + 1 header")


if __name__ == "__main__":
    main()
