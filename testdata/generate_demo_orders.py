#!/usr/bin/env python3
"""Generate demo Excel order files for the Bulk Offer Creator prototype."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADERS = [
    "Adress",
    "Aggregatbeteckning",
    "T-Flöde [l/s]",
    "Placering",
    "Aggregat",
    "Dimension på filterboxar i uteluftskanal",
]

DIM_COL_IDX = 6  # Column F (1-indexed)


def _write_sheet(ws, property_code, property_name, addresses, rows):
    """Write a single property sheet."""
    ws.title = f"{property_code} {property_name}"

    # Row 1: Title
    ws.cell(row=1, column=1, value=f"{property_code} {property_name}").font = TITLE_FONT

    # Row 3: Addresses
    ws.cell(row=3, column=1, value="Adresser:").font = Font(bold=True)
    ws.cell(row=3, column=2, value=addresses)

    # Row 5: Headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        if col_idx == DIM_COL_IDX:
            cell.fill = YELLOW_FILL

    # Data rows
    for row_idx, row_data in enumerate(rows, 6):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == DIM_COL_IDX:
                cell.fill = YELLOW_FILL

    # Auto-width columns
    for col_idx in range(1, len(HEADERS) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(5, 6 + len(rows))
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(max_len + 2, 12)


def generate_stockholmshem():
    """Generate the main demo file: 50 units across 6 properties."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: 0326 Kastanjen (8 units) ---
    ws = wb.active
    _write_sheet(ws, "0326", "Kastanjen",
        "Årstagatan 21 A-C, Kastanjegatan 2-10, Hjalmar Brantingsgatan 48 A-C",
        [
            ["Hus D, 26 lgh", "LB42", 1100, "Garage", "Envistar Flex", "800*600"],
            ["Hus D, 28 lgh", "LB41", 1100, "Garage", "Envistar Flex", "800*600"],
            ["Hus C, 36 lgh", "LB31", 1250, "Garage", "Envistar Flex", "800*600"],
            ["Hus B, 14 lgh", "LB21", 460, "Plan 1", "Envistar Flex", "600*300"],
            ["Hus A, 14 lgh", "LB11", 460, "Plan 1", "Envistar Flex", "600*300"],
            ["Hus A, 14 lgh", "LB12", 550, "Plan 1", "Envistar Flex", "600*400"],
            ["Hus B, 14 lgh", "LB22", 550, "Plan 1", "Envistar Flex", "600*400"],
            ["Hus C, 36 lgh", "LB32", 800, "Garage", "Envistar Flex", "600*600"],
        ]
    )

    # --- Sheet 2: 0331 Gröna Gatan 2 (10 units) ---
    ws2 = wb.create_sheet()
    _write_sheet(ws2, "0331", "Gröna Gatan 2",
        "Gröna Gatan 2A-E",
        [
            ["Gröna Gatan 2A", "LB01", 1400, "Garage", "Gold PX 030", "800*600"],
            ["Gröna Gatan 2A", "LB02", 1400, "Garage", "Gold PX 030", "800*600"],
            ["Gröna Gatan 2B", "LB03", 900, "Garage", "Gold PX 020", "600*600"],
            ["Gröna Gatan 2B", "LB04", 900, "Garage", "Gold PX 020", "600*600"],
            ["Gröna Gatan 2C", "LB05", 2000, "Garage", "Gold PX 050", "1000*600"],
            ["Gröna Gatan 2C", "LB06", 2000, "Garage", "Gold PX 050", "1000*600"],
            ["Gröna Gatan 2D", "LB07", 500, "Plan 1", "Envistar Flex", "600*300"],
            ["Gröna Gatan 2D", "LB08", 500, "Plan 1", "Envistar Flex", "600*300"],
            ["Gröna Gatan 2E", "LB09", 750, "Plan 1", "Envistar Flex", "600*400"],
            ["Gröna Gatan 2E", "LB10", 750, "Plan 1", "Envistar Flex", "600*400"],
        ]
    )

    # --- Sheet 3: 0333 Blåklinten (7 units) ---
    ws3 = wb.create_sheet()
    _write_sheet(ws3, "0333", "Blåklinten",
        "Verkmästargatan 27E-31, Mistelgatan 4A-6B",
        [
            ["Verkmästargatan 27E", "LB02", 1307, "Undercentral", "Gold PX 014", "1000*600"],
            ["Mistelgatan 4A-B", "LB03", 270, "Pannrum/Plan 1", "Gold PX 005", "400*300"],
            ["Verkmästargatan 29A", "LB04", 1500, "Undercentral", "Gold PX 020", "1000*600"],
            ["Verkmästargatan 29B", "LB05", 800, "Undercentral", "Gold PX 010", "600*500"],
            ["Mistelgatan 6A", "LB06", 350, "Pannrum", "Gold PX 005", "400*300"],
            ["Mistelgatan 6B", "LB07", 600, "Pannrum", "Gold PX 010", "600*400"],
            ["Verkmästargatan 31", "LB08", 1800, "Undercentral", "Gold PX 030", "800*600"],
        ]
    )

    # --- Sheet 4: 0334 Brillinge (9 units) ---
    ws4 = wb.create_sheet()
    _write_sheet(ws4, "0334", "Brillinge",
        "Brillingevägen 1A-9",
        [
            ["Brillingevägen 1A", "LB01", 1100, "Garage", "Envistar Flex", "800*600"],
            ["Brillingevägen 1B", "LB02", 1100, "Garage", "Envistar Flex", "800*600"],
            ["Brillingevägen 3A", "LB03", 1400, "Garage", "Envistar Flex", "800*600"],
            ["Brillingevägen 3B", "LB04", 460, "Plan 1", "Envistar Flex", "600*300"],
            ["Brillingevägen 5A", "LB05", 460, "Plan 1", "Envistar Flex", "600*300"],
            ["Brillingevägen 5B", "LB06", 800, "Plan 1", "Envistar Flex", "600*600"],
            ["Brillingevägen 7A", "LB07", 1600, "Garage", "Envistar Flex", "800*600"],
            ["Brillingevägen 7B", "LB08", 550, "Plan 1", "Envistar Flex", "600*400"],
            ["Brillingevägen 9", "LB09", 1800, "Garage", "Gold PX 030", "1000*600"],
        ]
    )

    # --- Sheet 5: 0412 Citrinen (8 units) ---
    ws5 = wb.create_sheet()
    _write_sheet(ws5, "0412", "Citrinen",
        "Citrongatan 1A-9",
        [
            ["Citrongatan 1A", "LB01", 2500, "Garage", "Gold PX 050", "1000*600"],
            ["Citrongatan 1B", "LB02", 2500, "Garage", "Gold PX 050", "1000*600"],
            ["Citrongatan 3", "LB03", 3000, "Garage", "Gold PX 080", "1400*700"],
            ["Citrongatan 5A", "LB04", 800, "Plan 1", "Envistar Flex", "600*400"],
            ["Citrongatan 5B", "LB05", 800, "Plan 1", "Envistar Flex", "600*400"],
            ["Citrongatan 7A", "LB06", 1100, "Garage", "Envistar Flex", "800*600"],
            ["Citrongatan 7B", "LB07", 5000, "Tak", "Gold PX 100", "1000*600"],
            ["Citrongatan 9", "LB08", 450, "Plan 1", "Envistar Flex", "600*300"],
        ]
    )

    # --- Sheet 6: 0550 August Söderman (8 units) ---
    ws6 = wb.create_sheet()
    _write_sheet(ws6, "0550", "August Söderman",
        "Södermangatan 1-7B",
        [
            ["Södermangatan 1", "LB01", 1200, "Undercentral", "Gold PX 020", "800*600"],
            ["Södermangatan 1", "LB02", 1200, "Undercentral", "Gold PX 020", "800*600"],
            ["Södermangatan 3A", "LB03", 700, "Undercentral", "Gold PX 010", "700*500"],
            ["Södermangatan 3B", "LB04", 700, "Undercentral", "Gold PX 010", "700*500"],
            ["Södermangatan 5", "LB05", 400, "Pannrum", "Envistar Flex", "600*300"],
            ["Södermangatan 5", "LB06", 400, "Pannrum", "Envistar Flex", "600*300"],
            ["Södermangatan 7A", "LB07", 1600, "Garage", "Gold PX 030", "1000*600"],
            ["Södermangatan 7B", "LB08", 550, "Pannrum", "Envistar Flex", "600*400"],
        ]
    )

    wb.save("testdata/stockholmshem_order.xlsx")
    print(f"Created stockholmshem_order.xlsx — {sum(ws.max_row - 5 for ws in wb.worksheets)} data rows across {len(wb.worksheets)} sheets")


def generate_quick_order():
    """Generate a simple demo file: 5 clean units, no ambiguities."""
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_sheet(ws, "0001", "Snabborder",
        "Exempelgatan 1A-5",
        [
            ["Exempelgatan 1A", "LB01", 1100, "Garage", "Envistar Flex", "800*600"],
            ["Exempelgatan 1B", "LB02", 1100, "Garage", "Envistar Flex", "800*600"],
            ["Exempelgatan 3A", "LB03", 900, "Garage", "Gold PX 020", "600*600"],
            ["Exempelgatan 3B", "LB04", 460, "Plan 1", "Envistar Flex", "600*300"],
            ["Exempelgatan 5", "LB05", 800, "Plan 1", "Envistar Flex", "600*600"],
        ]
    )

    wb.save("testdata/quick_order.xlsx")
    print("Created quick_order.xlsx — 5 data rows, 1 sheet")


if __name__ == "__main__":
    generate_stockholmshem()
    generate_quick_order()
