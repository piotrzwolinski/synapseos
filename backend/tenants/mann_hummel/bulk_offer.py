"""Bulk Offer Creator — AI-powered engine for batch processing client orders.

Parses Excel files with AHU specs, maps duct dimensions to housing variants + filters
via graph queries, and generates structured offers with Excel export.

Completely isolated from the main graph reasoning pipeline.
No Layer 4 sessions, no TechnicalState, no domain_config dependency.
"""

import io
import json
import os
import re
import uuid
import math
import time
import base64
import logging
from dataclasses import dataclass, field, asdict
from datetime import datetime
from typing import Optional, Generator

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google import genai
from google.genai import types
from db_result_helpers import result_to_dicts, result_single, result_value

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Gemini LLM client
# ---------------------------------------------------------------------------

_gemini_client = None
_MODEL = "gemini-2.0-flash"


def _get_gemini_client():
    global _gemini_client
    if _gemini_client is None:
        _gemini_client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))
    return _gemini_client

# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

@dataclass
class BulkOfferRow:
    row_id: int
    property_name: str
    address: str
    unit_id: str
    airflow_ls: float
    placement: str
    ahu_model: str
    duct_width: int
    duct_height: int
    # Populated after generation
    sheet_name: str = ""


@dataclass
class HousingMatch:
    variant_name: str          # e.g. "GDMI-900x600"
    product_code: str          # e.g. "GDMI-900x600-850-R-PG-AZ"
    width_mm: int
    height_mm: int
    housing_length: int
    material_code: str
    weight_kg: float = 0.0
    reference_airflow_m3h: float = 0.0
    modules_needed: int = 1


@dataclass
class FilterMatch:
    name: str                  # e.g. "Airpocket Eco ePM1 65% 592x592x635"
    model_name: str
    filter_class: str
    dimensions: str
    part_number: str
    slot_type: str             # "full", "half_width", "half_height"


@dataclass
class TransitionPiece:
    description: str           # e.g. "PT 900x600 - 800x600"
    housing_w: int
    housing_h: int
    duct_w: int
    duct_h: int


@dataclass
class GraphTrace:
    """Records which graph nodes/rules were consulted for a row."""
    nodes_consulted: list = field(default_factory=list)   # [{type, id, name, ...}]
    rules_applied: list = field(default_factory=list)     # [{rule, description}]
    reasoning_steps: list = field(default_factory=list)   # ["Step 1: ...", ...]


@dataclass
class OfferRowResult:
    row: BulkOfferRow
    housing: Optional[HousingMatch] = None
    filter_1: Optional[FilterMatch] = None
    filter_2: Optional[FilterMatch] = None
    transition: Optional[TransitionPiece] = None
    warnings: list = field(default_factory=list)
    error: Optional[str] = None
    graph_trace: Optional[GraphTrace] = None


@dataclass
class Clarification:
    id: str
    type: str
    severity: str              # "info", "warning", "critical"
    message: str
    options: list              # [{label, value, description}]
    affected_rows: list = field(default_factory=list)  # row_ids
    default_value: Optional[str] = None


@dataclass
class OfferConfig:
    material_code: str = "AZ"
    housing_length: int = 850
    filter_class: str = "ePM1 65%"
    product_family: str = "GDMI"
    overrides: dict = field(default_factory=dict)  # row_id -> {field: value}


@dataclass
class OfferSession:
    offer_id: str
    original_rows: list
    clarifications: list
    resolved_config: Optional[OfferConfig] = None
    results: list = field(default_factory=list)
    created_at: datetime = field(default_factory=datetime.now)
    llm_analysis: Optional[dict] = None
    filename: str = ""


# In-memory session store (sufficient for demo)
_offer_sessions: dict[str, OfferSession] = {}


# ---------------------------------------------------------------------------
# Column aliases for Swedish/English header detection
# ---------------------------------------------------------------------------

COLUMN_ALIASES = {
    "address":     ["adress", "address", "adr"],
    "unit_id":     ["aggregatbeteckning", "unit", "beteckning", "unit_id", "enhet"],
    "airflow_ls":  ["t-flöde [l/s]", "t-flöde", "airflow", "flöde", "airflow_ls", "l/s"],
    "placement":   ["placering", "placement", "location", "plats"],
    "ahu_model":   ["aggregat", "ahu", "ahu_model", "unit type", "typ"],
    "dimension":   ["dimension på filterboxar i uteluftskanal", "dimension", "dim",
                     "duct_width", "mått", "filterbox"],
}


def _match_column(header: str) -> Optional[str]:
    """Match a header string to a canonical column name."""
    h = header.strip().lower()
    for canonical, aliases in COLUMN_ALIASES.items():
        if any(alias in h for alias in aliases):
            return canonical
    return None


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def parse_excel(file_bytes: bytes) -> list[BulkOfferRow]:
    """Parse a multi-sheet Excel file into a flat list of order rows."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    rows = []
    row_id = 0

    for ws in wb.worksheets:
        sheet_name = ws.title
        # Extract property name from sheet title or row 1
        property_name = sheet_name

        # Find header row (look for a row containing recognizable headers)
        header_row = None
        col_map = {}
        for r in range(1, min(ws.max_row + 1, 15)):
            candidate = {}
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    matched = _match_column(val)
                    if matched:
                        candidate[matched] = c
            if len(candidate) >= 3:  # Need at least 3 recognized columns
                header_row = r
                col_map = candidate
                break

        if not header_row:
            continue  # Skip sheets without recognizable headers

        # Parse data rows
        for r in range(header_row + 1, ws.max_row + 1):
            # Skip blank rows
            first_val = ws.cell(row=r, column=1).value
            if first_val is None and ws.cell(row=r, column=2).value is None:
                continue

            address = str(ws.cell(row=r, column=col_map.get("address", 1)).value or "")
            unit_id = str(ws.cell(row=r, column=col_map.get("unit_id", 2)).value or "")
            placement = str(ws.cell(row=r, column=col_map.get("placement", 4)).value or "")
            ahu_model = str(ws.cell(row=r, column=col_map.get("ahu_model", 5)).value or "")

            # Parse airflow
            airflow_raw = ws.cell(row=r, column=col_map.get("airflow_ls", 3)).value
            try:
                airflow_ls = float(airflow_raw) if airflow_raw else 0.0
            except (ValueError, TypeError):
                airflow_ls = 0.0

            # Parse dimensions (format: "800*600" or "800x600")
            dim_raw = str(ws.cell(row=r, column=col_map.get("dimension", 6)).value or "")
            duct_w, duct_h = _parse_dimensions(dim_raw)

            if duct_w == 0 and duct_h == 0 and airflow_ls == 0:
                continue  # Skip empty rows

            row_id += 1
            rows.append(BulkOfferRow(
                row_id=row_id,
                property_name=property_name,
                address=address,
                unit_id=unit_id,
                airflow_ls=airflow_ls,
                placement=placement,
                ahu_model=ahu_model,
                duct_width=duct_w,
                duct_height=duct_h,
                sheet_name=sheet_name,
            ))

    return rows


def _parse_dimensions(dim_str: str) -> tuple[int, int]:
    """Parse dimension string like '800*600' or '800x600' into (width, height)."""
    m = re.match(r'(\d+)\s*[*xX×]\s*(\d+)', dim_str.strip())
    if m:
        return int(m.group(1)), int(m.group(2))
    return 0, 0


# ---------------------------------------------------------------------------
# Analysis & Clarification detection
# ---------------------------------------------------------------------------

def analyze_order(rows: list[BulkOfferRow], db) -> dict:
    """Analyze parsed rows and detect issues requiring clarification."""
    offer_id = str(uuid.uuid4())[:8]

    # Load available housing variants from graph
    variants = _load_housing_variants(db)
    capacity_rules = _load_capacity_rules(db)

    clarifications = []
    warnings = []
    stats = {
        "total_rows": len(rows),
        "properties": list(set(r.property_name for r in rows)),
        "property_count": len(set(r.property_name for r in rows)),
    }

    # 1. Global: Missing material
    clarifications.append(Clarification(
        id="material",
        type="MISSING_MATERIAL",
        severity="info",
        message=f"No material specified in the order. Which material for all {len(rows)} units?",
        options=[
            {"label": "Aluzink AZ (C4)", "value": "AZ", "description": "Standard outdoor-rated, corrosion class C4"},
            {"label": "Förzinkat FZ (C2)", "value": "FZ", "description": "Budget option, indoor only, C2"},
            {"label": "Zinkmagnesium ZM (C5)", "value": "ZM", "description": "Premium, highest corrosion resistance C5"},
        ],
        default_value="AZ",
    ))

    # 2. Global: Housing length
    clarifications.append(Clarification(
        id="housing_length",
        type="HOUSING_LENGTH",
        severity="info",
        message="Which housing length as default?",
        options=[
            {"label": "850mm (standard)", "value": "850", "description": "Standard depth, fits most filters up to 650mm"},
            {"label": "600mm (compact)", "value": "600", "description": "Compact depth, filters up to 450mm"},
        ],
        default_value="850",
    ))

    # 3. Global: Filter class
    clarifications.append(Clarification(
        id="filter_class",
        type="FILTER_CLASS",
        severity="info",
        message="Which filter class for the order?",
        options=[
            {"label": "ePM1 65% (standard)", "value": "ePM1 65%", "description": "Standard fine particulate filtration"},
            {"label": "ePM2.5 50%", "value": "ePM2.5 50%", "description": "Coarser filtration, lower pressure drop"},
            {"label": "ePM1 80%", "value": "ePM1 80%", "description": "Higher efficiency, premium"},
            {"label": "ePM1 90%", "value": "ePM1 90%", "description": "Near-HEPA efficiency, highest grade"},
        ],
        default_value="ePM1 65%",
    ))

    # 4. Per-row: Capacity issues
    capacity_warning_rows = []
    capacity_exceeded_rows = []
    for row in rows:
        airflow_m3h = row.airflow_ls * 3.6  # l/s → m³/h
        variant = _find_best_variant(row.duct_width, row.duct_height, variants)
        if variant:
            cap = _get_capacity(variant["width_mm"], variant["height_mm"], capacity_rules)
            if cap > 0:
                ratio = airflow_m3h / cap
                if ratio > 1.0:
                    capacity_exceeded_rows.append({
                        "row_id": row.row_id,
                        "unit_id": row.unit_id,
                        "property": row.property_name,
                        "airflow_ls": row.airflow_ls,
                        "airflow_m3h": airflow_m3h,
                        "capacity_m3h": cap,
                        "modules_needed": math.ceil(ratio),
                    })
                elif ratio > 0.85:
                    capacity_warning_rows.append({
                        "row_id": row.row_id,
                        "unit_id": row.unit_id,
                        "property": row.property_name,
                        "airflow_ls": row.airflow_ls,
                        "duct": f"{row.duct_width}x{row.duct_height}",
                    })

    if capacity_warning_rows:
        affected = [r["row_id"] for r in capacity_warning_rows]
        detail = ", ".join(f"{r['unit_id']} ({r['property']})" for r in capacity_warning_rows[:5])
        if len(capacity_warning_rows) > 5:
            detail += f" and {len(capacity_warning_rows) - 5} more"
        clarifications.append(Clarification(
            id="capacity_warning",
            type="CAPACITY_WARNING",
            severity="warning",
            message=f"{len(capacity_warning_rows)} units have airflow near module capacity limit (>85%): {detail}. Approve with margin notes or auto-upsize?",
            options=[
                {"label": "Approve with notes", "value": "approve", "description": "Keep current sizing, add warning notes"},
                {"label": "Auto-upsize all", "value": "upsize", "description": "Automatically select next larger housing"},
            ],
            affected_rows=affected,
            default_value="approve",
        ))

    if capacity_exceeded_rows:
        affected = [r["row_id"] for r in capacity_exceeded_rows]
        detail = ", ".join(
            f"{r['unit_id']} ({r['property']}, {r['airflow_ls']:.0f} l/s → {r['modules_needed']} modules)"
            for r in capacity_exceeded_rows[:5]
        )
        if len(capacity_exceeded_rows) > 5:
            detail += f" and {len(capacity_exceeded_rows) - 5} more"
        clarifications.append(Clarification(
            id="capacity_exceeded",
            type="CAPACITY_EXCEEDED",
            severity="critical",
            message=f"{len(capacity_exceeded_rows)} units exceed single module capacity: {detail}. Use multi-module configuration or override?",
            options=[
                {"label": "Auto multi-module", "value": "multi", "description": "Automatically stack modules where needed to handle capacity"},
                {"label": "Override all (accept overload)", "value": "override", "description": "Force single module with overload warnings for all"},
            ],
            affected_rows=affected,
            default_value="multi",
        ))

    # 5. Per-row: Non-standard dimensions
    for row in rows:
        variant = _find_best_variant(row.duct_width, row.duct_height, variants)
        if variant and (variant["width_mm"] - row.duct_width > 300 or variant["height_mm"] - row.duct_height > 300):
            clarifications.append(Clarification(
                id=f"no_match_{row.row_id}",
                type="NO_EXACT_MATCH",
                severity="warning",
                message=f"Unit {row.unit_id} ({row.property_name}): Duct {row.duct_width}x{row.duct_height} is far from nearest housing {variant['width_mm']}x{variant['height_mm']}. Large transition piece needed.",
                options=[
                    {"label": f"Use {variant['width_mm']}x{variant['height_mm']} + transition", "value": "accept", "description": "Accept oversized housing with transition piece"},
                    {"label": "Flag for manual review", "value": "manual", "description": "Skip this row, flag for engineer review"},
                ],
                affected_rows=[row.row_id],
                default_value="accept",
            ))

    # 6. Ambiguous placements
    ambiguous_placements = {"undercentral", "pannrum", "pannrum/plan 1"}
    ambiguous_rows = [r for r in rows if r.placement.lower() in ambiguous_placements]
    if ambiguous_rows:
        placements = list(set(r.placement for r in ambiguous_rows))
        affected = [r.row_id for r in ambiguous_rows]
        clarifications.append(Clarification(
            id="environment",
            type="MISSING_ENVIRONMENT",
            severity="info",
            message=f"{len(ambiguous_rows)} units have placement '{', '.join(placements)}'. Is this indoor or outdoor installation?",
            options=[
                {"label": "Indoor", "value": "indoor", "description": "Indoor environment — standard material sufficient"},
                {"label": "Outdoor", "value": "outdoor", "description": "Outdoor — requires insulated housing and higher corrosion class"},
            ],
            affected_rows=affected,
            default_value="indoor",
        ))

    # Store session
    session = OfferSession(
        offer_id=offer_id,
        original_rows=rows,
        clarifications=clarifications,
    )
    _offer_sessions[offer_id] = session

    return {
        "offer_id": offer_id,
        "stats": stats,
        "clarifications": [asdict(c) for c in clarifications],
        "row_count": len(rows),
        "properties": stats["properties"],
    }


# ---------------------------------------------------------------------------
# LLM-powered smart analysis (Gemini)
# ---------------------------------------------------------------------------

def llm_analyze_order(rows: list[BulkOfferRow], variants: list[dict],
                      capacity_rules: list[dict], filename: str = "") -> dict:
    """Use Gemini to produce an intelligent analysis of the uploaded order.

    Returns a dict with:
      - summary: str (1-line executive summary)
      - findings: list[{severity, units, message}] — flat, scannable list
    """
    # Build a compact table representation for the prompt
    rows_table = []
    for r in rows:
        rows_table.append({
            "id": r.row_id, "property": r.property_name, "unit": r.unit_id,
            "airflow_ls": r.airflow_ls, "duct": f"{r.duct_width}x{r.duct_height}",
            "placement": r.placement, "ahu": r.ahu_model,
        })

    # Build housing snap mapping so the LLM knows how ducts map to housings + capacity
    snap_examples = []
    for r in rows:
        best = _find_best_variant(r.duct_width, r.duct_height, variants)
        if best:
            cap = _get_capacity(best["width_mm"], best["height_mm"], capacity_rules)
            cap_str = f" (max {cap:.0f} m³/h = {cap/3.6:.0f} l/s)" if cap else ""
            snap_examples.append(f"  {r.duct_width}x{r.duct_height} → GDMI-{best['width_mm']}x{best['height_mm']}{cap_str}")
    snap_map = sorted(set(snap_examples))

    # Summarize available housing variants for context
    variant_sizes = sorted(set(f"{v['width_mm']}x{v['height_mm']}" for v in variants))
    cap_info = [f"  GDMI-{c['module_descriptor']}: max {c['output_rating']} m³/h" for c in capacity_rules]

    prompt = f"""You are a sales engineer at Mann+Hummel reviewing a client filter housing order.

## Order: {filename or "Client Order"}
{json.dumps(rows_table, indent=2, ensure_ascii=False)}

## IMPORTANT: How duct sizes map to GDMI housings
Client duct dimensions snap UP to the nearest available GDMI housing:
{chr(10).join(snap_map)}
Available GDMI sizes: {', '.join(variant_sizes)}
A duct of 800x600 is STANDARD — it maps to GDMI-900x600. Do NOT flag standard duct sizes as problems.

## Capacity rules (per housing module, in m³/h — client specifies l/s, multiply by 3.6)
{chr(10).join(cap_info)}

## Instructions
Return a flat JSON with ONE summary line and a list of findings. Each finding is max 15 words.
A sales engineer will scan this in 5 seconds — be telegraphic, not verbose.

{{
  "summary": "One sentence: X units across Y properties, [key characteristic].",
  "findings": [
    {{
      "severity": "action_required|review|info",
      "units": ["LB07"],
      "message": "5000 l/s exceeds single module — needs 3x GDMI-1200x600"
    }},
    {{
      "severity": "info",
      "units": [],
      "message": "All Kastanjen units are 800x600 — bulk opportunity"
    }}
  ]
}}

## Severity guide
- "action_required": Capacity exceeded, no housing match, impossible config. Engineer MUST act.
- "review": Near capacity limit (>85%), non-standard duct needing transition, mixed placements.
- "info": Bulk discount opportunity, standardization possible, placement pattern noted.

## Rules
- Max 6 findings. Only report things that matter.
- Each message: max 15 words. Start with the fact, not "The unit" or "Consider".
- Reference unit IDs and property names. Bold nothing — the UI handles formatting.
- Do NOT flag standard duct→housing mappings as problems (800x600→900x600 is normal).
- Do NOT flag transition pieces as problems — they are routine.
- DO flag: capacity exceeded, no housing match, very high airflow, placement ambiguity."""

    try:
        response = _get_gemini_client().models.generate_content(
            model=_MODEL,
            contents=[types.Content(role="user", parts=[types.Part.from_text(text=prompt)])],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.2,
                max_output_tokens=768,
            ),
        )
        result = json.loads(response.text)
        return result
    except Exception as e:
        logger.warning(f"LLM analysis failed: {e}")
        return {
            "summary": f"Parsed {len(rows)} unit(s) across {len(set(r.property_name for r in rows))} properties.",
            "findings": [],
        }


# ---------------------------------------------------------------------------
# Housing / filter lookup (graph queries)
# ---------------------------------------------------------------------------

_variants_cache: list[dict] | None = None
_capacity_cache: list[dict] | None = None


def _load_housing_variants(db) -> list[dict]:
    """Load all GDMI ProductVariant dimensions from graph."""
    global _variants_cache
    if _variants_cache:
        return _variants_cache

    graph = db.connect()
    result = graph.query("""
        MATCH (pv:ProductVariant)
        WHERE pv.product_family STARTS WITH 'FAM_GDMI'
        RETURN pv.name AS name,
               pv.width_mm AS width_mm,
               pv.height_mm AS height_mm,
               pv.product_family AS family,
               pv.reference_airflow_m3h AS airflow,
               pv.weight_kg AS weight_kg,
               pv.housing_length_mm AS housing_length
        ORDER BY pv.width_mm, pv.height_mm
    """)
    _variants_cache = result_to_dicts(result)
    return _variants_cache


def _load_capacity_rules(db) -> list[dict]:
    """Load capacity rules from graph."""
    global _capacity_cache
    if _capacity_cache:
        return _capacity_cache

    graph = db.connect()
    result = graph.query("""
        MATCH (cr:CapacityRule)
        WHERE cr.id STARTS WITH 'CAP_GDMI'
          AND NOT cr.id CONTAINS 'FLEX'
        RETURN cr.id AS id,
               cr.module_descriptor AS module_descriptor,
               cr.output_rating AS output_rating
    """)
    _capacity_cache = result_to_dicts(result)
    return _capacity_cache


def _find_best_variant(duct_w: int, duct_h: int, variants: list[dict]) -> Optional[dict]:
    """Find smallest GDMI variant that fits the duct dimensions."""
    # Filter to FAM_GDMI (not FLEX for standard orders)
    gdmi_variants = [v for v in variants if v["family"] == "FAM_GDMI"]

    # Find candidates where housing >= duct
    candidates = [v for v in gdmi_variants if v["width_mm"] >= duct_w and v["height_mm"] >= duct_h]

    if not candidates:
        # Try largest available
        if gdmi_variants:
            return max(gdmi_variants, key=lambda v: v["width_mm"] * v["height_mm"])
        return None

    # Return smallest fitting variant (minimize area)
    return min(candidates, key=lambda v: v["width_mm"] * v["height_mm"])


def _get_capacity(width_mm: int, height_mm: int, capacity_rules: list[dict]) -> float:
    """Get airflow capacity for a module size."""
    descriptor = f"{width_mm}x{height_mm}"
    # Try exact match on module descriptor
    for rule in capacity_rules:
        if rule["module_descriptor"] == descriptor:
            return float(rule["output_rating"])

    # For composite sizes, compute from 600x600 base modules
    base_cap = 3400.0  # GDMI 600x600 capacity
    h_modules = width_mm / 600
    v_modules = height_mm / 600
    return base_cap * h_modules * v_modules


def _load_filters_for_class(filter_class: str, db) -> dict:
    """Load demo filters grouped by slot type."""
    graph = db.connect()
    result = graph.query("""
        MATCH (fc:FilterConsumable {source: "BULK_OFFER_DEMO", filter_class: $fc})
        RETURN fc.name AS name,
               fc.model_name AS model_name,
               fc.filter_class AS filter_class,
               fc.dimensions AS dimensions,
               fc.part_number AS part_number,
               fc.module_width AS module_width,
               fc.module_height AS module_height
    """, fc=filter_class)
    filters = result_to_dicts(result)

    grouped = {"full": None, "half_width": None, "half_height": None}
    for f in filters:
        if f["module_width"] == 592 and f["module_height"] == 592:
            grouped["full"] = f
        elif f["module_width"] == 287 and f["module_height"] == 592:
            grouped["half_width"] = f
        elif f["module_width"] == 592 and f["module_height"] == 287:
            grouped["half_height"] = f
    return grouped


# ---------------------------------------------------------------------------
# Offer generation (streaming)
# ---------------------------------------------------------------------------

def generate_offer_streaming(offer_id: str, config: OfferConfig, db) -> Generator[dict, None, None]:
    """Process all rows and yield SSE events."""
    session = _offer_sessions.get(offer_id)
    if not session:
        yield {"type": "error", "detail": f"Offer session {offer_id} not found"}
        return

    session.resolved_config = config
    rows = session.original_rows
    variants = _load_housing_variants(db)
    capacity_rules = _load_capacity_rules(db)
    filters = _load_filters_for_class(config.filter_class, db)

    results = []
    housing_counts = {}
    current_property = None

    yield {"type": "start", "total": len(rows), "properties": list(set(r.property_name for r in rows))}

    for idx, row in enumerate(rows):
        # Property group header
        if row.property_name != current_property:
            if current_property:
                yield {"type": "property_done", "property": current_property}
            current_property = row.property_name
            yield {"type": "property_start", "property": current_property}

        # Build graph reasoning trace for this row
        trace = GraphTrace()

        # Find housing
        trace.reasoning_steps.append(f"Input: duct {row.duct_width}x{row.duct_height}, airflow {row.airflow_ls} l/s")
        variant = _find_best_variant(row.duct_width, row.duct_height, variants)
        result = OfferRowResult(row=row, graph_trace=trace)

        if not variant:
            result.error = f"No housing variant found for {row.duct_width}x{row.duct_height}"
            trace.reasoning_steps.append(f"FAILED: No ProductVariant >= {row.duct_width}x{row.duct_height}")
            results.append(result)
            yield {"type": "row_result", "row_id": row.row_id, "status": "error",
                   "detail": result.error, "row": idx + 1, "total": len(rows)}
            continue

        # Record variant selection in trace
        trace.nodes_consulted.append({
            "type": "ProductVariant", "id": variant.get("name", ""),
            "detail": f"{variant['width_mm']}x{variant['height_mm']}mm"
        })
        trace.reasoning_steps.append(
            f"Matched duct {row.duct_width}x{row.duct_height} → ProductVariant {variant['width_mm']}x{variant['height_mm']}"
        )

        # Compute modules needed
        airflow_m3h = row.airflow_ls * 3.6
        cap = _get_capacity(variant["width_mm"], variant["height_mm"], capacity_rules)
        modules_needed = max(1, math.ceil(airflow_m3h / cap)) if cap > 0 else 1

        # Record capacity rule in trace
        trace.nodes_consulted.append({
            "type": "CapacityRule",
            "id": f"CAP_GDMI_{variant['width_mm']}x{variant['height_mm']}",
            "detail": f"max {cap:.0f} m³/h"
        })
        trace.rules_applied.append({
            "rule": "Capacity Check",
            "description": f"{airflow_m3h:.0f} m³/h vs {cap:.0f} m³/h capacity → {modules_needed} module(s)"
        })
        trace.reasoning_steps.append(
            f"Capacity: {airflow_m3h:.0f}/{cap:.0f} m³/h ({airflow_m3h/cap*100:.0f}%) → {modules_needed} module(s)"
        )

        # Check overrides from clarification answers
        override = config.overrides.get(str(row.row_id), {})
        if override.get("capacity") == "override":
            modules_needed = 1
            trace.rules_applied.append({"rule": "User Override", "description": "Forced single module"})

        # Build product code
        product_code = f"GDMI-{variant['width_mm']}x{variant['height_mm']}-{config.housing_length}-R-PG-{config.material_code}"
        trace.reasoning_steps.append(f"Product code: {product_code}")

        # Weight lookup
        weight_key = f"weight_kg_{config.housing_length}" if config.housing_length in (600, 850) else "weight_kg"
        weight = variant.get(weight_key, variant.get("weight_kg", 0)) or 0

        housing = HousingMatch(
            variant_name=f"GDMI-{variant['width_mm']}x{variant['height_mm']}",
            product_code=product_code,
            width_mm=variant["width_mm"],
            height_mm=variant["height_mm"],
            housing_length=config.housing_length,
            material_code=config.material_code,
            weight_kg=float(weight) * modules_needed,
            reference_airflow_m3h=float(variant.get("airflow", 0) or 0),
            modules_needed=modules_needed,
        )
        result.housing = housing

        # Count
        key = housing.variant_name
        housing_counts[key] = housing_counts.get(key, 0) + modules_needed

        # Filter selection
        if filters.get("full"):
            f = filters["full"]
            result.filter_1 = FilterMatch(
                name=f["name"], model_name=f["model_name"],
                filter_class=f["filter_class"], dimensions=f["dimensions"],
                part_number=f["part_number"], slot_type="full",
            )
            trace.nodes_consulted.append({
                "type": "FilterConsumable", "id": f["part_number"],
                "detail": f["name"]
            })
            trace.reasoning_steps.append(f"Filter: {f['name']} (592x592 full module)")

        # Half-module filter (Filter 2) for housings with half modules
        has_half = variant["height_mm"] % 600 != 0 or variant["width_mm"] % 600 != 0
        if has_half:
            half_w = filters.get("half_width")
            half_h = filters.get("half_height")
            if variant["width_mm"] < 600 and half_w:
                result.filter_2 = FilterMatch(
                    name=half_w["name"], model_name=half_w["model_name"],
                    filter_class=half_w["filter_class"], dimensions=half_w["dimensions"],
                    part_number=half_w["part_number"], slot_type="half_width",
                )
                trace.nodes_consulted.append({
                    "type": "FilterConsumable", "id": half_w["part_number"],
                    "detail": half_w["name"]
                })
            elif variant["height_mm"] < 600 and half_h:
                result.filter_2 = FilterMatch(
                    name=half_h["name"], model_name=half_h["model_name"],
                    filter_class=half_h["filter_class"], dimensions=half_h["dimensions"],
                    part_number=half_h["part_number"], slot_type="half_height",
                )
                trace.nodes_consulted.append({
                    "type": "FilterConsumable", "id": half_h["part_number"],
                    "detail": half_h["name"]
                })

        # Transition piece
        if row.duct_width != variant["width_mm"] or row.duct_height != variant["height_mm"]:
            result.transition = TransitionPiece(
                description=f"PT {variant['width_mm']}x{variant['height_mm']} - {row.duct_width}x{row.duct_height}",
                housing_w=variant["width_mm"],
                housing_h=variant["height_mm"],
                duct_w=row.duct_width,
                duct_h=row.duct_height,
            )
            trace.rules_applied.append({
                "rule": "Transition Required",
                "description": f"Duct {row.duct_width}x{row.duct_height} ≠ Housing {variant['width_mm']}x{variant['height_mm']}"
            })
            trace.reasoning_steps.append(
                f"Transition: PT {variant['width_mm']}x{variant['height_mm']} → {row.duct_width}x{row.duct_height}"
            )

        # Warnings
        if cap > 0:
            ratio = airflow_m3h / cap
            if ratio > 0.85 and modules_needed == 1:
                result.warnings.append(f"Airflow at {ratio*100:.0f}% of capacity ({airflow_m3h:.0f}/{cap:.0f} m³/h)")

        results.append(result)

        # Yield progress with graph trace
        yield {
            "type": "row_result",
            "row_id": row.row_id,
            "row": idx + 1,
            "total": len(rows),
            "status": "success",
            "property": row.property_name,
            "unit_id": row.unit_id,
            "duct": f"{row.duct_width}x{row.duct_height}",
            "housing": product_code,
            "filter_1": result.filter_1.name if result.filter_1 else None,
            "filter_2": result.filter_2.name if result.filter_2 else None,
            "transition": result.transition.description if result.transition else None,
            "modules_needed": modules_needed,
            "warnings": result.warnings,
            "graph_trace": {
                "nodes_consulted": trace.nodes_consulted,
                "rules_applied": trace.rules_applied,
                "reasoning_steps": trace.reasoning_steps,
            },
        }

        # Small delay for streaming effect in demo
        time.sleep(0.05)

    # Final property done
    if current_property:
        yield {"type": "property_done", "property": current_property}

    # Store results
    session.results = results

    # Summary
    yield {
        "type": "summary",
        "offer_id": offer_id,
        "total": len(rows),
        "success": len([r for r in results if r.housing]),
        "errors": len([r for r in results if r.error]),
        "housing_counts": housing_counts,
        "properties": list(set(r.row.property_name for r in results)),
    }

    yield {"type": "complete", "offer_id": offer_id}


# ---------------------------------------------------------------------------
# Natural language refinement (LLM-powered)
# ---------------------------------------------------------------------------

def llm_interpret_refinement(user_message: str, offer_id: str, db) -> dict:
    """Use Gemini to interpret a natural language refinement request.

    Returns a dict with:
      - interpretation: str (what the system understood)
      - changes: dict[row_id_str, {field: new_value}] (to apply via refine endpoint)
      - affected_count: int
      - requires_regeneration: bool
    """
    session = _offer_sessions.get(offer_id)
    if not session:
        return {"interpretation": "Offer session not found.", "changes": {}, "affected_count": 0, "requires_regeneration": False}

    # Build context of current offer state
    rows_context = []
    for r in session.original_rows:
        rows_context.append({
            "row_id": r.row_id, "property": r.property_name, "unit_id": r.unit_id,
            "duct_w": r.duct_width, "duct_h": r.duct_height,
            "airflow_ls": r.airflow_ls, "placement": r.placement,
        })

    results_context = []
    for r in session.results:
        results_context.append({
            "row_id": r.row.row_id, "unit_id": r.row.unit_id, "property": r.row.property_name,
            "housing": r.housing.product_code if r.housing else None,
            "warnings": r.warnings,
        })

    config_context = {}
    if session.resolved_config:
        config_context = {
            "material": session.resolved_config.material_code,
            "housing_length": session.resolved_config.housing_length,
            "filter_class": session.resolved_config.filter_class,
        }

    prompt = f"""You are an HVAC offer assistant. The user wants to modify a generated offer.

## Current offer config
{json.dumps(config_context, indent=2)}

## Current rows (input)
{json.dumps(rows_context, indent=2, ensure_ascii=False)}

## Current results
{json.dumps(results_context[:20], indent=2, ensure_ascii=False)}
{"..." if len(results_context) > 20 else ""}

## User's request
"{user_message}"

## Instructions
Interpret the user's request and determine which changes to apply. Return JSON:
{{
  "interpretation": "Human-readable explanation of what you understood",
  "changes": {{
    "<row_id>": {{"duct_width": 600, "duct_height": 600}},
    "<row_id>": {{"airflow_ls": 500}}
  }},
  "config_changes": {{
    "material_code": "ZM",
    "housing_length": 600,
    "filter_class": "ePM1 80%"
  }},
  "affected_count": 5,
  "requires_regeneration": true
}}

Rules:
- "changes" maps row_id (as string) to field changes. ONLY these exact field names: "duct_width", "duct_height", "airflow_ls" (NOT duct_w, NOT duct_h)
- "config_changes" for global settings (material_code, housing_length, filter_class). Omit fields that don't change.
- If user says "change all 600x300 to 600x600", find all rows with those duct dimensions and change them
- If user mentions a property name, only affect rows in that property
- If user mentions specific unit IDs, only affect those units
- requires_regeneration = true if any changes were made
- Be generous in interpretation — match partial property names, unit IDs, etc."""

    try:
        response = _get_gemini_client().models.generate_content(
            model=_MODEL,
            contents=[types.Content(role="user", parts=[types.Part.from_text(text=prompt)])],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.1,
                max_output_tokens=1024,
            ),
        )
        result = json.loads(response.text)
        return result
    except Exception as e:
        logger.warning(f"LLM refinement interpretation failed: {e}")
        return {
            "interpretation": f"I couldn't interpret your request: {user_message}",
            "changes": {},
            "config_changes": {},
            "affected_count": 0,
            "requires_regeneration": False,
        }


# ---------------------------------------------------------------------------
# Email drafting (LLM-powered)
# ---------------------------------------------------------------------------

def draft_offer_email(offer_id: str, language: str = "sv") -> dict:
    """Use Gemini to draft a professional email for the generated offer.

    Returns a dict with:
      - subject: str
      - body: str
      - language: str
    """
    session = _offer_sessions.get(offer_id)
    if not session or not session.results:
        return {"subject": "", "body": "No offer data available.", "language": language}

    # Summarize the offer
    properties = list(set(r.row.property_name for r in session.results))
    total_units = len(session.results)
    successful = len([r for r in session.results if r.housing])
    housing_counts = {}
    for r in session.results:
        if r.housing:
            key = r.housing.variant_name
            housing_counts[key] = housing_counts.get(key, 0) + r.housing.modules_needed

    config = session.resolved_config
    material_name = {"AZ": "Aluzink (C4)", "FZ": "Förzinkat (C2)", "ZM": "Zinkmagnesium (C5)"}.get(
        config.material_code if config else "AZ", config.material_code if config else "AZ"
    )

    lang_instruction = "Write in Swedish (formal business Swedish)." if language == "sv" else "Write in English (formal business English)."

    prompt = f"""You are a senior sales engineer at Mann+Hummel writing an email to a client about their filter housing order.

## Offer details
- Client order: {session.filename or "Client order"}
- Properties: {', '.join(properties)}
- Total units: {total_units} ({successful} successfully configured)
- Material: {material_name}
- Housing length: {config.housing_length if config else 850}mm
- Filter class: {config.filter_class if config else 'ePM1 65%'}

## Housing breakdown
{chr(10).join(f"- {name}: {count} st" for name, count in sorted(housing_counts.items()))}

## Warnings/notes
{len([r for r in session.results if r.warnings])} units have capacity margin warnings.
{len([r for r in session.results if r.transition])} units need transition pieces.

## Instructions
{lang_instruction}

Write a professional email with:
1. Subject line
2. Greeting
3. Brief intro referencing the client's order
4. Summary of what's included
5. Any important notes (capacity warnings, transition pieces)
6. Next steps (request for confirmation, delivery timeline discussion)
7. Professional sign-off

Return JSON: {{"subject": "...", "body": "..."}}

The tone should be professional but warm — this is a proposal, not a contract."""

    try:
        response = _get_gemini_client().models.generate_content(
            model=_MODEL,
            contents=[types.Content(role="user", parts=[types.Part.from_text(text=prompt)])],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.5,
                max_output_tokens=1024,
            ),
        )
        result = json.loads(response.text)
        result["language"] = language
        return result
    except Exception as e:
        logger.warning(f"Email drafting failed: {e}")
        return {"subject": "Offer", "body": f"Error generating email: {e}", "language": language}


# ---------------------------------------------------------------------------
# PDF parsing (Gemini multimodal)
# ---------------------------------------------------------------------------

def parse_pdf_order(file_bytes: bytes, filename: str = "") -> list[BulkOfferRow]:
    """Use Gemini multimodal to extract order data from a PDF file.

    Sends the PDF to Gemini and asks it to extract the tabular data
    in the same format as the Excel parser expects.
    """
    b64_data = base64.b64encode(file_bytes).decode("utf-8")

    prompt = """You are an HVAC order document parser. Extract the tabular order data from this PDF.

The PDF contains a filter housing order from a client. Extract ALL rows of data into a structured format.

Return JSON array:
[
  {
    "property_name": "Property or building name",
    "address": "Street address",
    "unit_id": "Unit/aggregate identifier (e.g., LB01)",
    "airflow_ls": 1100,
    "placement": "Where the unit is placed (e.g., Garage, Plan 1, Tak)",
    "ahu_model": "AHU model name (e.g., Envistar Flex, Gold PX 020)",
    "duct_width": 800,
    "duct_height": 600
  }
]

Rules:
- Extract dimensions from any format: "800x600", "800*600", "Ø400", etc.
- For circular ducts (Ø), convert to equivalent rectangular: width=diameter, height=diameter
- Airflow may be in l/s or m³/h — always convert to l/s (divide m³/h by 3.6)
- If the PDF has multiple pages or sections for different properties, capture the property name
- Include ALL rows, even if some fields are partially empty
- For Swedish text: "Flöde" = airflow, "Placering" = placement, "Aggregat" = AHU
- If no property name is found, use the filename or "Imported from PDF"
"""

    try:
        response = _get_gemini_client().models.generate_content(
            model=_MODEL,
            contents=[types.Content(
                role="user",
                parts=[
                    types.Part(inline_data=types.Blob(mime_type="application/pdf", data=b64_data)),
                    types.Part.from_text(text=prompt),
                ],
            )],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.0,
                max_output_tokens=4096,
            ),
        )

        raw_rows = json.loads(response.text)
        rows = []
        for idx, raw in enumerate(raw_rows, 1):
            rows.append(BulkOfferRow(
                row_id=idx,
                property_name=raw.get("property_name", filename or "PDF Import"),
                address=raw.get("address", ""),
                unit_id=raw.get("unit_id", f"U{idx:02d}"),
                airflow_ls=float(raw.get("airflow_ls", 0)),
                placement=raw.get("placement", ""),
                ahu_model=raw.get("ahu_model", ""),
                duct_width=int(raw.get("duct_width", 0)),
                duct_height=int(raw.get("duct_height", 0)),
                sheet_name=raw.get("property_name", "Sheet1"),
            ))
        return rows

    except Exception as e:
        logger.warning(f"PDF parsing failed: {e}")
        raise ValueError(f"Failed to extract data from PDF: {e}")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

EXPORT_HEADERS = [
    "Dimension på filterboxar i uteluftskanal",
    "Filterskåp",
    "Art",
    "Filter 1",
    "Filter 2",
    "Övergångar",
    "Moduler",
    "Varningar",
]


def generate_offer_excel(offer_id: str) -> Optional[bytes]:
    """Generate Excel output matching Mann+Hummel format."""
    session = _offer_sessions.get(offer_id)
    if not session or not session.results:
        return None

    wb = openpyxl.Workbook()
    results = session.results

    # Group by property
    by_property = {}
    for r in results:
        prop = r.row.property_name
        if prop not in by_property:
            by_property[prop] = []
        by_property[prop].append(r)

    first_sheet = True
    for prop_name, prop_results in by_property.items():
        if first_sheet:
            ws = wb.active
            first_sheet = False
        else:
            ws = wb.create_sheet()
        ws.title = prop_name[:31]  # Excel sheet name limit

        # Title
        ws.cell(row=1, column=1, value=prop_name).font = TITLE_FONT

        # Input headers (row 4)
        input_headers = ["Adress", "Aggregatbeteckning", "T-Flöde [l/s]", "Placering", "Aggregat",
                         "Dimension"]
        for col, h in enumerate(input_headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            if col == 6:
                cell.fill = YELLOW_FILL

        # Output headers (continuing columns)
        output_start = len(input_headers) + 1
        for col_offset, h in enumerate(EXPORT_HEADERS):
            col = output_start + col_offset
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.fill = GREEN_FILL

        # Data rows
        for row_idx, r in enumerate(prop_results, 5):
            row_data = [
                r.row.address,
                r.row.unit_id,
                r.row.airflow_ls,
                r.row.placement,
                r.row.ahu_model,
                f"{r.row.duct_width}*{r.row.duct_height}",
            ]
            # Output columns
            if r.housing:
                row_data.extend([
                    r.housing.product_code,
                    "",  # Art number (placeholder)
                    r.filter_1.name if r.filter_1 else "",
                    r.filter_2.name if r.filter_2 else "",
                    r.transition.description if r.transition else "",
                    r.housing.modules_needed if r.housing.modules_needed > 1 else "",
                    "; ".join(r.warnings) if r.warnings else "",
                ])
            else:
                row_data.extend([r.error or "No match", "", "", "", "", "", ""])

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = THIN_BORDER
                if col == 6:
                    cell.fill = YELLOW_FILL
                if r.warnings and col >= output_start:
                    cell.fill = ORANGE_FILL

        # Auto-width
        for col in range(1, output_start + len(EXPORT_HEADERS) + 1):
            max_len = max(
                (len(str(ws.cell(row=r, column=col).value or "")) for r in range(4, 5 + len(prop_results))),
                default=10,
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max(max_len + 2, 10), 40)

    # --- Sammanställning (Summary) sheet ---
    ws_sum = wb.create_sheet("Sammanställning")
    ws_sum.cell(row=1, column=1, value="Sammanställning").font = TITLE_FONT

    # Housing counts
    ws_sum.cell(row=3, column=1, value="Filterskåp").font = HEADER_FONT
    ws_sum.cell(row=3, column=2, value="Antal").font = HEADER_FONT
    housing_counts = {}
    for r in results:
        if r.housing:
            key = r.housing.product_code
            housing_counts[key] = housing_counts.get(key, 0) + r.housing.modules_needed
    for idx, (code, count) in enumerate(sorted(housing_counts.items()), 4):
        ws_sum.cell(row=idx, column=1, value=code).border = THIN_BORDER
        ws_sum.cell(row=idx, column=2, value=count).border = THIN_BORDER

    total_row = 4 + len(housing_counts)
    ws_sum.cell(row=total_row, column=1, value="Totalsumma").font = HEADER_FONT
    ws_sum.cell(row=total_row, column=2, value=sum(housing_counts.values())).font = HEADER_FONT

    # Filter counts
    filter_start = total_row + 2
    ws_sum.cell(row=filter_start, column=1, value="Filter").font = HEADER_FONT
    ws_sum.cell(row=filter_start, column=2, value="Antal").font = HEADER_FONT
    filter_counts = {}
    for r in results:
        if r.filter_1:
            filter_counts[r.filter_1.name] = filter_counts.get(r.filter_1.name, 0) + 1
        if r.filter_2:
            filter_counts[r.filter_2.name] = filter_counts.get(r.filter_2.name, 0) + 1
    for idx, (name, count) in enumerate(sorted(filter_counts.items()), filter_start + 1):
        ws_sum.cell(row=idx, column=1, value=name).border = THIN_BORDER
        ws_sum.cell(row=idx, column=2, value=count).border = THIN_BORDER

    ws_sum.column_dimensions["A"].width = 45
    ws_sum.column_dimensions["B"].width = 12

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ===========================================================================
# COMPETITOR CROSS-REFERENCE MODE
# ===========================================================================

@dataclass
class CompetitorItem:
    """A single line item extracted from a competitor document."""
    line_id: int
    raw_text: str
    competitor_manufacturer: str
    competitor_model: str
    competitor_code: str
    category: str                    # bag_filter, compact_filter, panel_filter, hepa_filter, housing
    iso_class: str
    width_mm: int
    height_mm: int
    depth_mm: int
    quantity: int
    application_context: str = ""


@dataclass
class CrossRefResult:
    """Result of cross-referencing a single competitor item."""
    competitor: CompetitorItem
    mh_product_name: str
    mh_product_code: str
    mh_housing_family: str
    confidence: float
    match_type: str                  # graph_exact, graph_near, llm_inferred, no_match
    dimension_note: str
    performance_note: str
    graph_trace: Optional[GraphTrace] = None


@dataclass
class CrossRefSession:
    """Session for a cross-reference operation."""
    offer_id: str
    mode: str = "cross_reference"
    competitor_items: list = field(default_factory=list)
    cross_ref_results: list = field(default_factory=list)
    created_at: datetime = field(default_factory=datetime.now)
    filename: str = ""
    llm_analysis: Optional[dict] = None
    # After generation, these mirror OfferSession fields
    original_rows: list = field(default_factory=list)
    results: list = field(default_factory=list)
    resolved_config: Optional[OfferConfig] = None
    clarifications: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Competitor context loading (cached)
# ---------------------------------------------------------------------------

_competitor_cache: list[dict] | None = None


def _load_competitor_context(db) -> list[dict]:
    """Load CompetitorProduct nodes from graph for LLM context and matching."""
    global _competitor_cache
    if _competitor_cache:
        return _competitor_cache

    graph = db.connect()
    result = graph.query("""
        MATCH (cp:CompetitorProduct)
        OPTIONAL MATCH (cp)-[r:CROSS_REFERENCES]->(target)
        RETURN cp.id AS id, cp.manufacturer AS manufacturer,
               cp.product_line AS product_line, cp.model AS model,
               cp.category AS category, cp.iso_class AS iso_class,
               cp.width_mm AS width_mm, cp.height_mm AS height_mm,
               cp.depth_mm AS depth_mm, cp.aliases AS aliases,
               collect(CASE WHEN target IS NOT NULL THEN {
                 target_name: COALESCE(target.name, target.id),
                 target_code: target.part_number,
                 confidence: r.confidence,
                 match_type: r.match_type,
                 dimension_note: r.dimension_note,
                 performance_note: r.performance_note
               } END) AS mappings
    """)
    _competitor_cache = result_to_dicts(result)
    return _competitor_cache


def _load_all_mh_filters(db) -> list[dict]:
    """Load all MH filter consumables for LLM context."""
    graph = db.connect()
    result = graph.query("""
        MATCH (fc:FilterConsumable)
        WHERE fc.source = 'BULK_OFFER_DEMO'
        RETURN fc.name AS name, fc.model_name AS model_name,
               fc.filter_class AS filter_class, fc.dimensions AS dimensions,
               fc.part_number AS part_number, fc.filter_type AS filter_type
    """)
    return result_to_dicts(result)


# ---------------------------------------------------------------------------
# Competitor document parsing (LLM multimodal)
# ---------------------------------------------------------------------------

def parse_competitor_document(file_bytes: bytes, filename: str, db) -> list[CompetitorItem]:
    """Use Gemini multimodal to extract competitor product items from any document."""
    competitor_context = _load_competitor_context(db)

    # Build context summary for the prompt
    known_products = []
    for cp in competitor_context:
        entry = f"- {cp['manufacturer']} {cp['model']} ({cp['category']}) {cp['iso_class'] or ''}"
        if cp['aliases']:
            entry += f" aliases: {', '.join(cp['aliases'][:4])}"
        known_products.append(entry)

    prompt = f"""You are an HVAC filter expert. Extract ALL competitor filter product line items from this document.
The document may be an order form, email, PDF catalog, scanned document, or free-form text.

## Known competitor products for reference:
{chr(10).join(known_products)}

## Instructions
- Detect the competitor manufacturer (likely Camfil, but could be others)
- Extract every distinct product with its specifications
- Parse dimensions in mm (width x height x depth)
- Map filter classes to ISO 16890 (ePM1, ePM2.5, ePM10) or legacy EN 779 (F7, F9, H13)
- If quantity is not specified, default to 1
- For housing products (CamCube, CamBox), set category to "housing"
- Handle Swedish, English, German text

Return a JSON array. Each element:
{{
  "raw_text": "original text snippet for this item",
  "competitor_manufacturer": "Camfil",
  "competitor_model": "Hi-Flo XLS",
  "competitor_code": "part number or empty string",
  "category": "bag_filter|compact_filter|panel_filter|hepa_filter|housing",
  "iso_class": "ePM1 80%",
  "width_mm": 592,
  "height_mm": 592,
  "depth_mm": 635,
  "quantity": 12,
  "application_context": "ventilation|cleanroom|industrial|unknown"
}}

Rules:
- Always return an array, even if only 1 item
- Convert EN 779 classes: F7=ePM1 65%, F8=ePM1 80%, F9=ePM1 90%, G4=Coarse 75%, M5=ePM2.5 50%
- If dimensions are given as "592x592-8x600", that means 592x592mm face, 600mm depth, 8 pockets
- If unsure about a field, make your best guess and note it in raw_text
"""

    # Determine MIME type
    fname = filename.lower()
    if fname.endswith(".pdf"):
        mime = "application/pdf"
    elif fname.endswith((".png", ".jpg", ".jpeg")):
        mime = "image/png" if fname.endswith(".png") else "image/jpeg"
    elif fname.endswith((".xlsx", ".xls")):
        # For Excel, extract text first then send as text
        return _parse_competitor_excel(file_bytes, filename, db)
    else:
        mime = "application/octet-stream"

    b64_data = base64.b64encode(file_bytes).decode("utf-8")

    try:
        response = _get_gemini_client().models.generate_content(
            model=_MODEL,
            contents=[types.Content(
                role="user",
                parts=[
                    types.Part(inline_data=types.Blob(mime_type=mime, data=b64_data)),
                    types.Part.from_text(text=prompt),
                ],
            )],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.0,
                max_output_tokens=4096,
            ),
        )

        raw_items = json.loads(response.text)
        return _raw_to_competitor_items(raw_items)

    except Exception as e:
        logger.warning(f"Competitor document parsing failed: {e}")
        raise ValueError(f"Failed to extract competitor products: {e}")


def _parse_competitor_excel(file_bytes: bytes, filename: str, db) -> list[CompetitorItem]:
    """Parse Excel file by extracting cell text and sending to LLM."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    text_lines = []
    for ws in wb.worksheets:
        text_lines.append(f"--- Sheet: {ws.title} ---")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), values_only=True):
            vals = [str(v) for v in row if v is not None]
            if vals:
                text_lines.append(" | ".join(vals))

    text_content = "\n".join(text_lines)

    competitor_context = _load_competitor_context(db)
    known_products = []
    for cp in competitor_context:
        entry = f"- {cp['manufacturer']} {cp['model']} ({cp['category']}) {cp['iso_class'] or ''}"
        known_products.append(entry)

    prompt = f"""You are an HVAC filter expert. Extract ALL competitor filter product line items from this spreadsheet data.

## Known competitor products:
{chr(10).join(known_products)}

## Spreadsheet content:
{text_content}

Return a JSON array of competitor products found. Each element:
{{
  "raw_text": "original text for this item",
  "competitor_manufacturer": "Camfil",
  "competitor_model": "Hi-Flo XLS",
  "competitor_code": "part number or empty string",
  "category": "bag_filter|compact_filter|panel_filter|hepa_filter|housing",
  "iso_class": "ePM1 80%",
  "width_mm": 592, "height_mm": 592, "depth_mm": 635,
  "quantity": 12,
  "application_context": "ventilation|cleanroom|industrial|unknown"
}}

Convert EN 779: F7=ePM1 65%, F8=ePM1 80%, F9=ePM1 90%, G4=Coarse 75%, M5=ePM2.5 50%.
If no products found, return empty array []."""

    try:
        response = _get_gemini_client().models.generate_content(
            model=_MODEL,
            contents=[types.Content(role="user", parts=[types.Part.from_text(text=prompt)])],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.0,
                max_output_tokens=4096,
            ),
        )
        raw_items = json.loads(response.text)
        return _raw_to_competitor_items(raw_items)
    except Exception as e:
        logger.warning(f"Competitor Excel parsing failed: {e}")
        raise ValueError(f"Failed to extract competitor products from Excel: {e}")


def _raw_to_competitor_items(raw_items: list[dict]) -> list[CompetitorItem]:
    """Convert raw LLM output dicts to CompetitorItem dataclasses."""
    items = []
    for idx, raw in enumerate(raw_items, 1):
        items.append(CompetitorItem(
            line_id=idx,
            raw_text=raw.get("raw_text") or "",
            competitor_manufacturer=raw.get("competitor_manufacturer") or "Unknown",
            competitor_model=raw.get("competitor_model") or "Unknown",
            competitor_code=raw.get("competitor_code") or "",
            category=raw.get("category") or "unknown",
            iso_class=raw.get("iso_class") or "",
            width_mm=int(raw.get("width_mm") or 0),
            height_mm=int(raw.get("height_mm") or 0),
            depth_mm=int(raw.get("depth_mm") or 0),
            quantity=int(raw.get("quantity") or 1),
            application_context=raw.get("application_context") or "unknown",
        ))
    return items


# ---------------------------------------------------------------------------
# Hybrid matching: Graph lookup + LLM fallback
# ---------------------------------------------------------------------------

def match_competitor_items(items: list[CompetitorItem], db) -> list[CrossRefResult]:
    """Match competitor items to MH products using graph + LLM hybrid."""
    results = []

    for item in items:
        trace = GraphTrace()
        trace.reasoning_steps.append(
            f"Input: {item.competitor_manufacturer} {item.competitor_model} "
            f"{item.iso_class} {item.width_mm}x{item.height_mm}x{item.depth_mm}"
        )

        # Step 1: Graph lookup (high confidence)
        graph_match = _graph_lookup_competitor(item, db, trace)

        if graph_match and graph_match.get("confidence", 0) >= 0.7:
            results.append(CrossRefResult(
                competitor=item,
                mh_product_name=graph_match.get("target_name", ""),
                mh_product_code=graph_match.get("part_number", ""),
                mh_housing_family=_infer_housing_family(graph_match),
                confidence=graph_match["confidence"],
                match_type="graph_exact" if graph_match["confidence"] >= 0.9 else "graph_near",
                dimension_note=graph_match.get("dimension_note", ""),
                performance_note=graph_match.get("performance_note", ""),
                graph_trace=trace,
            ))
            continue

        # Step 2: Fuzzy graph match by category + dimensions
        fuzzy_match = _graph_fuzzy_lookup(item, db, trace)
        if fuzzy_match and fuzzy_match.get("confidence", 0) >= 0.6:
            results.append(CrossRefResult(
                competitor=item,
                mh_product_name=fuzzy_match.get("target_name", ""),
                mh_product_code=fuzzy_match.get("part_number", ""),
                mh_housing_family=_infer_housing_family(fuzzy_match),
                confidence=fuzzy_match["confidence"],
                match_type="graph_near",
                dimension_note=fuzzy_match.get("dimension_note", ""),
                performance_note=fuzzy_match.get("performance_note", ""),
                graph_trace=trace,
            ))
            continue

        # Step 3: LLM fallback
        llm_match = _llm_match_competitor(item, db, trace)
        results.append(CrossRefResult(
            competitor=item,
            mh_product_name=llm_match.get("mh_product_name", ""),
            mh_product_code=llm_match.get("mh_part_number", ""),
            mh_housing_family=llm_match.get("mh_housing_family", ""),
            confidence=llm_match.get("confidence", 0),
            match_type=llm_match.get("match_type", "llm_inferred"),
            dimension_note=llm_match.get("dimension_note", ""),
            performance_note=llm_match.get("performance_note", ""),
            graph_trace=trace,
        ))

    return results


def _graph_lookup_competitor(item: CompetitorItem, db, trace: GraphTrace) -> Optional[dict]:
    """Try exact model + class match in the CompetitorProduct graph."""
    graph = db.connect()
    result = graph.query("""
        MATCH (cp:CompetitorProduct)-[r:CROSS_REFERENCES]->(target)
        WHERE cp.manufacturer = $manufacturer
          AND (cp.model = $model
               OR $model IN cp.aliases
               OR ANY(alias IN cp.aliases WHERE toLower(alias) = toLower($model)))
          AND (cp.iso_class = $iso_class OR $iso_class = '' OR cp.iso_class IS NULL)
        RETURN cp.model AS competitor_model,
               labels(target)[0] AS target_type,
               COALESCE(target.name, target.id) AS target_name,
               target.part_number AS part_number,
               r.confidence AS confidence,
               r.match_type AS match_type,
               r.dimension_note AS dimension_note,
               r.performance_note AS performance_note
        ORDER BY r.confidence DESC
        LIMIT 1
    """, manufacturer=item.competitor_manufacturer,
         model=item.competitor_model,
         iso_class=item.iso_class or "")

    record = result_single(result)
    if record:
        d = dict(record)
        trace.nodes_consulted.append({
            "type": "CompetitorProduct",
            "id": item.competitor_model,
            "detail": f"Graph match → {d['target_name']}"
        })
        trace.reasoning_steps.append(
            f"Graph match: {d['competitor_model']} → {d['target_name']} "
            f"(confidence: {d['confidence']:.0%})"
        )
        trace.rules_applied.append({
            "rule": "Graph Cross-Reference",
            "description": f"{d['match_type']}: {d.get('dimension_note', '')}"
        })
        return d
    return None


def _graph_fuzzy_lookup(item: CompetitorItem, db, trace: GraphTrace) -> Optional[dict]:
    """Fuzzy match by category + approximate dimensions."""
    graph = db.connect()
    result = graph.query("""
        MATCH (cp:CompetitorProduct)-[r:CROSS_REFERENCES]->(target)
        WHERE cp.manufacturer = $manufacturer
          AND cp.category = $category
          AND abs(cp.width_mm - $w) <= 20
          AND abs(cp.height_mm - $h) <= 20
        RETURN cp.model AS competitor_model,
               COALESCE(target.name, target.id) AS target_name,
               target.part_number AS part_number,
               r.confidence AS confidence,
               r.match_type AS match_type,
               r.dimension_note AS dimension_note,
               r.performance_note AS performance_note
        ORDER BY r.confidence DESC
        LIMIT 1
    """, manufacturer=item.competitor_manufacturer,
         category=item.category,
         w=item.width_mm, h=item.height_mm)

    record = result_single(result)
    if record:
        d = dict(record)
        trace.reasoning_steps.append(
            f"Fuzzy graph: category={item.category}, dims ~{item.width_mm}x{item.height_mm} "
            f"→ {d['target_name']} (confidence: {d['confidence']:.0%})"
        )
        return d
    return None


def _llm_match_competitor(item: CompetitorItem, db, trace: GraphTrace) -> dict:
    """Use Gemini to infer the best MH equivalent for an unknown competitor product."""
    mh_filters = _load_all_mh_filters(db)

    prompt = f"""You are an HVAC filter cross-reference expert at Mann+Hummel.

A client has specified this competitor product:
- Manufacturer: {item.competitor_manufacturer}
- Model: {item.competitor_model}
- Code: {item.competitor_code}
- Category: {item.category}
- ISO Class: {item.iso_class}
- Dimensions: {item.width_mm}x{item.height_mm}x{item.depth_mm}mm
- Application: {item.application_context}

Available Mann+Hummel filter products:
{json.dumps(mh_filters, indent=2, ensure_ascii=False)}

Find the best MH equivalent. Return JSON:
{{
  "mh_product_name": "product name from the list above",
  "mh_part_number": "part number",
  "mh_housing_family": "GDMI or GDB or GDP or GDC",
  "confidence": 0.85,
  "match_type": "llm_inferred",
  "dimension_note": "any dimension differences",
  "performance_note": "any performance differences",
  "reasoning": "brief explanation"
}}

Confidence guide:
- 0.9+: exact type + class + dimension match
- 0.7-0.9: same type + class, different dimensions
- 0.5-0.7: same type, different class
- <0.5: no good match (set match_type to "no_match")

Housing families: GDMI (compact/modular), GDB (bag filter), GDP (panel), GDC (carbon).
For HEPA: 610x610 Camfil → 592x592 MH (different standards, adapter needed)."""

    try:
        response = _get_gemini_client().models.generate_content(
            model=_MODEL,
            contents=[types.Content(role="user", parts=[types.Part.from_text(text=prompt)])],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.1,
                max_output_tokens=512,
            ),
        )
        result = json.loads(response.text)

        trace.reasoning_steps.append(
            f"LLM inferred: {item.competitor_model} → {result.get('mh_product_name', '?')} "
            f"(confidence: {result.get('confidence', 0):.0%})"
        )
        trace.rules_applied.append({
            "rule": "LLM Cross-Reference",
            "description": result.get("reasoning", "")
        })

        if result.get("confidence", 0) < 0.5:
            result["match_type"] = "no_match"
        return result

    except Exception as e:
        logger.warning(f"LLM competitor matching failed: {e}")
        trace.reasoning_steps.append(f"LLM fallback failed: {e}")
        return {
            "mh_product_name": "", "mh_part_number": "",
            "mh_housing_family": "", "confidence": 0,
            "match_type": "no_match", "dimension_note": "",
            "performance_note": "", "reasoning": f"LLM error: {e}",
        }


def _infer_housing_family(match: dict) -> str:
    """Infer MH housing family from match result."""
    name = (match.get("target_name") or "").lower()
    if "gdmi" in name or "aircube" in name or "nanoclass" in name:
        return "GDMI"
    if "gdb" in name or "airpocket" in name:
        return "GDB"
    if "gdp" in name or "airpanel" in name or "airsquare" in name:
        return "GDP"
    if "gdc" in name or "eco-c" in name:
        return "GDC"
    return ""


# ---------------------------------------------------------------------------
# Competitor order analysis
# ---------------------------------------------------------------------------

def analyze_competitor_order(items: list[CompetitorItem], filename: str, db) -> dict:
    """Analyze parsed competitor items, run cross-reference matching."""
    offer_id = str(uuid.uuid4())[:8]

    # Run the matching
    cross_ref_results = match_competitor_items(items, db)

    # Detect clarifications
    clarifications = []

    low_confidence = [r for r in cross_ref_results if 0 < r.confidence < 0.7]
    no_matches = [r for r in cross_ref_results if r.match_type == "no_match"]

    # Material clarification (same as standard mode)
    clarifications.append(Clarification(
        id="material",
        type="MISSING_MATERIAL",
        severity="info",
        message=f"Which material for the MH housings? ({len(items)} units)",
        options=[
            {"label": "Aluzink AZ (C4)", "value": "AZ", "description": "Standard outdoor-rated"},
            {"label": "Förzinkat FZ (C2)", "value": "FZ", "description": "Budget, indoor only"},
            {"label": "Zinkmagnesium ZM (C5)", "value": "ZM", "description": "Premium, highest corrosion resistance"},
        ],
        default_value="AZ",
    ))

    # Housing length
    clarifications.append(Clarification(
        id="housing_length",
        type="HOUSING_LENGTH",
        severity="info",
        message="Default MH housing length?",
        options=[
            {"label": "850mm (standard)", "value": "850", "description": "Fits most filters up to 650mm"},
            {"label": "600mm (compact)", "value": "600", "description": "Filters up to 450mm"},
        ],
        default_value="850",
    ))

    if no_matches:
        affected = [r.competitor.line_id for r in no_matches]
        detail = ", ".join(f"{r.competitor.competitor_model}" for r in no_matches[:5])
        clarifications.append(Clarification(
            id="no_match_items",
            type="NO_MATCH",
            severity="critical",
            message=f"{len(no_matches)} items have no MH equivalent: {detail}",
            options=[
                {"label": "Skip these items", "value": "skip", "description": "Exclude from offer"},
                {"label": "Use closest match", "value": "closest", "description": "Use best available alternative"},
            ],
            affected_rows=affected,
            default_value="skip",
        ))

    if low_confidence:
        affected = [r.competitor.line_id for r in low_confidence]
        clarifications.append(Clarification(
            id="low_confidence",
            type="LOW_CONFIDENCE_MATCH",
            severity="warning",
            message=f"{len(low_confidence)} items have uncertain matches (<70% confidence). Accept best guesses?",
            options=[
                {"label": "Accept all", "value": "accept", "description": "Use best available match"},
                {"label": "Skip uncertain", "value": "skip", "description": "Only include high-confidence matches"},
            ],
            affected_rows=affected,
            default_value="accept",
        ))

    # Store session
    session = CrossRefSession(
        offer_id=offer_id,
        competitor_items=items,
        cross_ref_results=cross_ref_results,
        filename=filename,
        clarifications=[asdict(c) for c in clarifications],
    )
    _offer_sessions[offer_id] = session

    # Stats
    stats = {
        "total_items": len(items),
        "matched": len([r for r in cross_ref_results if r.match_type != "no_match"]),
        "high_confidence": len([r for r in cross_ref_results if r.confidence >= 0.8]),
        "low_confidence": len(low_confidence),
        "no_match": len(no_matches),
        "manufacturers": list(set(i.competitor_manufacturer for i in items)),
    }

    return {
        "offer_id": offer_id,
        "mode": "cross_reference",
        "stats": stats,
        "row_count": len(items),
        "properties": stats["manufacturers"],
        "cross_ref_results": [_crossref_to_dict(r) for r in cross_ref_results],
        "clarifications": [asdict(c) for c in clarifications],
    }


def _crossref_to_dict(r: CrossRefResult) -> dict:
    """Serialize a CrossRefResult for JSON response."""
    return {
        "line_id": r.competitor.line_id,
        "competitor": f"{r.competitor.competitor_manufacturer} {r.competitor.competitor_model}",
        "competitor_code": r.competitor.competitor_code,
        "competitor_dims": f"{r.competitor.width_mm}x{r.competitor.height_mm}x{r.competitor.depth_mm}",
        "quantity": r.competitor.quantity,
        "mh_product": r.mh_product_name,
        "mh_code": r.mh_product_code,
        "mh_housing_family": r.mh_housing_family,
        "confidence": r.confidence,
        "match_type": r.match_type,
        "dimension_note": r.dimension_note,
        "performance_note": r.performance_note,
        "graph_trace": asdict(r.graph_trace) if r.graph_trace else None,
    }


def llm_analyze_crossref(items: list[CompetitorItem],
                          crossref_results: list[dict],
                          filename: str = "") -> dict:
    """Gemini analysis of the cross-reference results."""
    items_summary = [
        {"model": i.competitor_model, "class": i.iso_class,
         "dims": f"{i.width_mm}x{i.height_mm}x{i.depth_mm}", "qty": i.quantity}
        for i in items
    ]

    prompt = f"""You are a sales engineer at Mann+Hummel reviewing a competitor cross-reference.

## Competitor Order: {filename}
{json.dumps(items_summary, indent=2, ensure_ascii=False)}

## Cross-Reference Results (first 20)
{json.dumps(crossref_results[:20], indent=2, ensure_ascii=False)}

Return JSON:
{{
  "summary": "One sentence: X Camfil items mapped to MH equivalents, Y high-confidence, Z need review.",
  "findings": [
    {{
      "severity": "action_required|review|info",
      "units": ["Hi-Flo XLS x12"],
      "message": "max 15 words about this finding"
    }}
  ]
}}

Max 6 findings. Each max 15 words. Focus on:
- Dimension mismatches (esp. 610→592 for HEPA)
- No-match items
- Efficiency upgrades/downgrades
- Bulk quantities (opportunity signals)
- Do NOT flag routine mappings as problems."""

    try:
        response = _get_gemini_client().models.generate_content(
            model=_MODEL,
            contents=[types.Content(role="user", parts=[types.Part.from_text(text=prompt)])],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.2,
                max_output_tokens=768,
            ),
        )
        return json.loads(response.text)
    except Exception as e:
        logger.warning(f"LLM crossref analysis failed: {e}")
        return {
            "summary": f"Parsed {len(items)} competitor items from {filename}.",
            "findings": [],
        }


# ---------------------------------------------------------------------------
# Cross-ref offer generation (streaming)
# ---------------------------------------------------------------------------

def generate_crossref_offer_streaming(
    offer_id: str, config: OfferConfig, db
) -> Generator[dict, None, None]:
    """Convert cross-ref results to BulkOfferRows and generate MH offer via existing pipeline."""
    session = _offer_sessions.get(offer_id)
    if not session or not hasattr(session, "cross_ref_results"):
        yield {"type": "error", "detail": f"Cross-ref session {offer_id} not found"}
        return

    cross_refs = session.cross_ref_results

    # Phase 1: Yield cross-reference mapping events
    yield {"type": "crossref_start", "total": len(cross_refs)}
    for cr in cross_refs:
        yield {
            "type": "crossref_mapping",
            "line_id": cr.competitor.line_id,
            "competitor": f"{cr.competitor.competitor_manufacturer} {cr.competitor.competitor_model}",
            "mh_product": cr.mh_product_name,
            "confidence": cr.confidence,
            "match_type": cr.match_type,
            "dimension_note": cr.dimension_note,
            "graph_trace": asdict(cr.graph_trace) if cr.graph_trace else None,
        }
        time.sleep(0.03)
    yield {"type": "crossref_complete"}

    # Phase 2: Convert to BulkOfferRows for the standard pipeline
    rows = []
    for cr in cross_refs:
        if cr.match_type == "no_match":
            continue
        # Infer duct dimensions from the MH housing family
        # Use competitor dims as duct size (the standard pipeline will snap to nearest housing)
        w = cr.competitor.width_mm or 600
        h = cr.competitor.height_mm or 600
        for q in range(cr.competitor.quantity):
            rows.append(BulkOfferRow(
                row_id=len(rows) + 1,
                property_name=cr.competitor.competitor_manufacturer,
                address="",
                unit_id=f"{cr.competitor.competitor_model} #{q+1}" if cr.competitor.quantity > 1 else cr.competitor.competitor_model,
                airflow_ls=0,  # Not applicable for competitor cross-ref
                placement="",
                ahu_model=cr.competitor.competitor_code,
                duct_width=w,
                duct_height=h,
                sheet_name=cr.competitor.competitor_manufacturer,
            ))

    if not rows:
        yield {"type": "error", "detail": "No matchable items to generate offer for"}
        return

    # Store as regular session for the standard pipeline
    session.original_rows = rows
    session.resolved_config = config

    # Phase 3: Delegate to existing generation pipeline
    for event in generate_offer_streaming(offer_id, config, db):
        yield event


def llm_interpret_crossref_refinement(user_message: str, offer_id: str, db) -> dict:
    """LLM-powered refinement for cross-ref results."""
    session = _offer_sessions.get(offer_id)
    if not session:
        return {"interpretation": "Session not found.", "requires_regeneration": False}

    # Build context from cross-ref results
    crossref_context = []
    if hasattr(session, "cross_ref_results") and session.cross_ref_results:
        for cr in session.cross_ref_results[:30]:
            crossref_context.append({
                "line_id": cr.competitor.line_id,
                "competitor": f"{cr.competitor.competitor_manufacturer} {cr.competitor.competitor_model}",
                "mh_product": cr.mh_product_name,
                "confidence": cr.confidence,
                "match_type": cr.match_type,
            })

    prompt = f"""You are a sales engineer processing a competitor cross-reference offer.

## Current cross-reference mappings:
{json.dumps(crossref_context, indent=2, ensure_ascii=False)}

## User message:
"{user_message}"

Interpret what the user wants to change. Return JSON:
{{
  "interpretation": "User-friendly explanation of what you understood",
  "changes": {{}},
  "config_changes": {{}},
  "affected_count": 0,
  "requires_regeneration": true
}}

Possible changes:
- config_changes: {{"material_code": "ZM"}}, {{"housing_length": 600}}, {{"filter_class": "ePM1 80%"}}
- If user asks to override a specific mapping, note it in interpretation (manual override not yet supported)
"""

    try:
        response = _get_gemini_client().models.generate_content(
            model=_MODEL,
            contents=[types.Content(role="user", parts=[types.Part.from_text(text=prompt)])],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                temperature=0.1,
                max_output_tokens=512,
            ),
        )
        return json.loads(response.text)
    except Exception as e:
        logger.warning(f"LLM crossref refinement failed: {e}")
        return {
            "interpretation": f"Could not interpret: {e}",
            "requires_regeneration": False,
        }
